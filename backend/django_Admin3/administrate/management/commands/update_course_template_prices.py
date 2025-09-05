import logging
import os
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.db import transaction
from administrate.models import CourseTemplate, PriceLevel, CourseTemplatePriceLevel

logger = logging.getLogger(__name__)

class Command(BaseCommand):
    help = 'Update course template price levels from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Path to the Excel file containing price updates'
        )
        parser.add_argument(
            '--debug',
            action='store_true',
            help='Enable debug logging'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Run without making actual changes to the database'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='Name of the Excel sheet to read (default: first sheet)'
        )
        parser.add_argument(
            '--start-row',
            type=int,
            default=1,
            help='Row number to start reading data (default: 1, assuming headers in row 0)'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        debug = options['debug']
        dry_run = options['dry_run']
        sheet_name = options['sheet']
        start_row = options['start_row']
        
        if debug:
            logger.setLevel(logging.DEBUG)
        
        # Check if file exists
        if not os.path.exists(excel_file):
            self.stdout.write(
                self.style.ERROR(f'Excel file not found: {excel_file}')
            )
            return
        
        # Try to import pandas, fall back to openpyxl if not available
        try:
            import pandas as pd
            use_pandas = True
        except ImportError:
            try:
                import openpyxl
                use_pandas = False
            except ImportError:
                self.stdout.write(
                    self.style.ERROR(
                        'Please install either pandas or openpyxl to read Excel files:\n'
                        'pip install pandas\n'
                        'or\n'
                        'pip install openpyxl'
                    )
                )
                return
        
        if dry_run:
            self.stdout.write(
                self.style.WARNING('DRY RUN MODE - No changes will be made to the database')
            )
        
        try:
            if use_pandas:
                self._process_with_pandas(
                    excel_file, sheet_name, start_row, debug, dry_run
                )
            else:
                self._process_with_openpyxl(
                    excel_file, sheet_name, start_row, debug, dry_run
                )
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Error processing Excel file: {str(e)}')
            )
            if debug:
                logger.exception(e)

    def _process_with_pandas(self, excel_file, sheet_name, start_row, debug, dry_run):
        """Process Excel file using pandas"""
        import pandas as pd
        
        # Read the Excel file
        self.stdout.write(f'Reading Excel file: {excel_file}')
        
        if sheet_name:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=start_row-1)
        else:
            df = pd.read_excel(excel_file, header=start_row-1)
        
        self.stdout.write(f'Found {len(df)} rows in the Excel file')
        
        # Get all price levels and course templates for lookup
        price_levels = {pl.name: pl for pl in PriceLevel.objects.all()}
        course_templates = {ct.code: ct for ct in CourseTemplate.objects.all()}
        
        updated_count = 0
        skipped_count = 0
        error_count = 0
        not_found_count = 0
        multiple_records_count = 0
        
        # Process each row
        for index, row in df.iterrows():
            row_num = index + start_row
            
            # Get values from columns B, C, and D
            # Assuming column B is at index 1, C at index 2, D at index 3
            price_level_name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else None
            amount_str = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else None
            course_code = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else None
            
            if debug:
                self.stdout.write(
                    f'Row {row_num}: Price Level: {price_level_name}, '
                    f'Amount: {amount_str}, Course Code: {course_code}'
                )
            
            # Skip rows with missing data
            if not all([price_level_name, amount_str, course_code]):
                skipped_count += 1
                if debug:
                    self.stdout.write(
                        self.style.WARNING(f'Row {row_num}: Skipping due to missing data')
                    )
                continue
            
            # Process the row
            result = self._update_price(
                price_level_name, amount_str, course_code,
                price_levels, course_templates,
                row_num, debug, dry_run
            )
            
            if result == 'updated':
                updated_count += 1
            elif result == 'not_found':
                not_found_count += 1
            elif result == 'multiple':
                multiple_records_count += 1
            elif result == 'error':
                error_count += 1
            else:
                skipped_count += 1
        
        # Print summary
        self._print_summary(
            updated_count, skipped_count, error_count,
            not_found_count, multiple_records_count, dry_run
        )

    def _process_with_openpyxl(self, excel_file, sheet_name, start_row, debug, dry_run):
        """Process Excel file using openpyxl"""
        import openpyxl
        
        # Read the Excel file
        self.stdout.write(f'Reading Excel file: {excel_file}')
        
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                self.stdout.write(
                    self.style.ERROR(f'Sheet "{sheet_name}" not found in Excel file')
                )
                return
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Get all price levels and course templates for lookup
        price_levels = {pl.name: pl for pl in PriceLevel.objects.all()}
        course_templates = {ct.code: ct for ct in CourseTemplate.objects.all()}
        
        updated_count = 0
        skipped_count = 0
        error_count = 0
        not_found_count = 0
        multiple_records_count = 0
        
        # Process each row starting from start_row
        for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            # Get values from columns B, C, and D (indices 1, 2, 3)
            price_level_name = str(row[1]).strip() if row[1] else None
            amount_str = str(row[2]).strip() if row[2] else None
            course_code = str(row[3]).strip() if row[3] else None
            
            if debug:
                self.stdout.write(
                    f'Row {row_num}: Price Level: {price_level_name}, '
                    f'Amount: {amount_str}, Course Code: {course_code}'
                )
            
            # Skip rows with missing data
            if not all([price_level_name, amount_str, course_code]):
                skipped_count += 1
                if debug:
                    self.stdout.write(
                        self.style.WARNING(f'Row {row_num}: Skipping due to missing data')
                    )
                continue
            
            # Process the row
            result = self._update_price(
                price_level_name, amount_str, course_code,
                price_levels, course_templates,
                row_num, debug, dry_run
            )
            
            if result == 'updated':
                updated_count += 1
            elif result == 'not_found':
                not_found_count += 1
            elif result == 'multiple':
                multiple_records_count += 1
            elif result == 'error':
                error_count += 1
            else:
                skipped_count += 1
        
        wb.close()
        
        # Print summary
        self._print_summary(
            updated_count, skipped_count, error_count,
            not_found_count, multiple_records_count, dry_run
        )

    def _update_price(self, price_level_name, amount_str, course_code,
                     price_levels, course_templates,
                     row_num, debug, dry_run):
        """Update a single price entry"""
        
        # Map "Repeat / Additional" to "Retaker"
        if price_level_name == "Repeat / Additional":
            price_level_name = "Retaker"
            if debug:
                self.stdout.write(
                    f'Row {row_num}: Mapping "Repeat / Additional" to "Retaker"'
                )
        
        # Check if price level exists
        if price_level_name not in price_levels:
            self.stdout.write(
                self.style.WARNING(
                    f'Row {row_num}: Price level "{price_level_name}" not found in database'
                )
            )
            return 'not_found'
        
        # Check if course template exists
        if course_code not in course_templates:
            self.stdout.write(
                self.style.WARNING(
                    f'Row {row_num}: Course template "{course_code}" not found in database'
                )
            )
            return 'not_found'
        
        # Parse amount
        try:
            # Remove any currency symbols and commas
            amount_str = amount_str.replace('£', '').replace('$', '').replace(',', '').strip()
            amount = Decimal(amount_str)
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(
                    f'Row {row_num}: Invalid amount format "{amount_str}": {str(e)}'
                )
            )
            return 'error'
        
        price_level = price_levels[price_level_name]
        course_template = course_templates[course_code]
        
        # Get CourseTemplatePriceLevel records
        price_level_records = CourseTemplatePriceLevel.objects.filter(
            course_template=course_template,
            price_level=price_level
        )
        
        # Check if there's exactly one record
        if price_level_records.count() == 0:
            self.stdout.write(
                self.style.WARNING(
                    f'Row {row_num}: No price level record found for '
                    f'{course_code} - {price_level_name}'
                )
            )
            return 'not_found'
        elif price_level_records.count() > 1:
            self.stdout.write(
                self.style.WARNING(
                    f'Row {row_num}: Multiple price level records found for '
                    f'{course_code} - {price_level_name} ({price_level_records.count()} records). '
                    f'Skipping update.'
                )
            )
            return 'multiple'
        
        # Update the price
        price_record = price_level_records.first()
        old_amount = price_record.amount
        
        if old_amount == amount:
            if debug:
                self.stdout.write(
                    f'Row {row_num}: Price unchanged for {course_code} - {price_level_name}: £{amount}'
                )
            return 'unchanged'
        
        if not dry_run:
            try:
                with transaction.atomic():
                    price_record.amount = amount
                    price_record.save()
                    
                    self.stdout.write(
                        self.style.SUCCESS(
                            f'Row {row_num}: Updated {course_code} - {price_level_name}: '
                            f'£{old_amount} -> £{amount}'
                        )
                    )
            except Exception as e:
                self.stdout.write(
                    self.style.ERROR(
                        f'Row {row_num}: Error updating price: {str(e)}'
                    )
                )
                return 'error'
        else:
            self.stdout.write(
                f'Row {row_num}: [DRY RUN] Would update {course_code} - {price_level_name}: '
                f'£{old_amount} -> £{amount}'
            )
        
        return 'updated'

    def _print_summary(self, updated_count, skipped_count, error_count,
                      not_found_count, multiple_records_count, dry_run):
        """Print summary of the operation"""
        
        mode = '[DRY RUN] ' if dry_run else ''
        
        self.stdout.write(
            self.style.SUCCESS(
                f'\n{mode}Price update completed:\n'
                f'  Updated: {updated_count}\n'
                f'  Not found: {not_found_count}\n'
                f'  Multiple records (skipped): {multiple_records_count}\n'
                f'  Skipped (missing data): {skipped_count}\n'
                f'  Errors: {error_count}'
            )
        )