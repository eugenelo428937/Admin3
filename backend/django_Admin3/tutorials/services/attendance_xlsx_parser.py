"""Parse the instructor-uploaded attendance xlsx.

Defensive parsing: read-only + data_only to disable formula evaluation;
reject cells whose raw value starts with '=' (formula injection); silently
skip rows with a blank Attendance cell; reject student_refs not enrolled in
the given session.

The parser is decoupled from any HTTP / view code so it can be unit-tested
in isolation. The view (Task 17) is responsible for the 2 MB size cap and
magic-byte check before calling this parser.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import IO, Dict, List

from openpyxl import load_workbook

VALID_STATUSES = {'ATTENDED', 'ABSENT', 'LATE', 'OTHER'}
HEADER_COLS = ('Title', 'First Name', 'Last Name', 'Student Ref', 'Email', 'Company', 'Attendance')


@dataclass
class ParseRow:
    student_ref: int
    status: str
    reason: str = ''


@dataclass
class ParseResult:
    rows: List[ParseRow] = field(default_factory=list)
    skipped_blank: int = 0
    errors: List[str] = field(default_factory=list)
    ref_to_registration_id: Dict[int, int] = field(default_factory=dict)


def parse_attendance_xlsx(file: IO[bytes], session) -> ParseResult:
    """Return a ParseResult summarising what would be written.

    Caller (the upload view) is responsible for converting ``rows`` into the
    ``items`` shape expected by ``save_attendance_items`` using
    ``ref_to_registration_id`` for the join.
    """
    # Local import to avoid a model-import cycle at service-module load time.
    from tutorials.models import TutorialRegistration

    # data_only=False so formulas surface as their raw '=...' strings; the
    # startswith('=') guard below then rejects them explicitly. With
    # data_only=True, openpyxl would return a cached calculation result
    # (or None for unsaved workbooks), masking malicious formula content.
    wb = load_workbook(file, read_only=True, data_only=False)
    ws = wb.active

    registration_pairs = list(
        TutorialRegistration.objects
        .filter(tutorial_session=session)
        .values_list('student__student_ref', 'id')
    )
    valid_refs = {ref for ref, _ in registration_pairs}
    ref_to_reg_id = {ref: rid for ref, rid in registration_pairs}

    result = ParseResult(ref_to_registration_id=ref_to_reg_id)
    rows_iter = ws.iter_rows(
        min_row=2, max_col=len(HEADER_COLS), values_only=False,
    )
    for row_num, row_cells in enumerate(rows_iter, start=2):
        if all(c.value in (None, '') for c in row_cells):
            continue
        _title, _fn, _ln, ref_cell, _email, _company, status_cell = row_cells

        status_val = status_cell.value
        if status_val in (None, ''):
            result.skipped_blank += 1
            continue

        # Formula injection guard. read_only + data_only prevents evaluation,
        # but a string literal beginning with '=' would still survive.
        if isinstance(status_val, str) and status_val.startswith('='):
            result.errors.append(f'row {row_num}: formula not allowed in Attendance cell')
            continue
        if not isinstance(status_val, str):
            result.errors.append(f'row {row_num}: Attendance must be text')
            continue
        status = status_val.strip().upper()
        if status not in VALID_STATUSES:
            result.errors.append(f'row {row_num}: invalid status {status_val!r}')
            continue

        ref_raw = ref_cell.value
        try:
            ref = int(ref_raw)
        except (TypeError, ValueError):
            result.errors.append(
                f'row {row_num}: Student Ref must be an integer (got {ref_raw!r})'
            )
            continue
        if ref not in valid_refs:
            result.errors.append(
                f'row {row_num}: student_ref {ref} not enrolled in this session'
            )
            continue

        result.rows.append(ParseRow(student_ref=ref, status=status))

    return result
