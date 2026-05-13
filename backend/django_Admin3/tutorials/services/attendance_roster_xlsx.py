"""Generate an attendance roster xlsx for a TutorialSession.

Schema (sheet 1): Title | First Name | Last Name | Student Ref | Email | Company | Attendance
- Student Ref is the join key on upload.
- Company is intentionally blank (placeholder until the data source exists).
- Attendance column carries a data-validation dropdown.

Sheet 2 ("Meta") records session_id, event_code, session_date, generated_at —
informational only; not parsed on upload.
"""
from __future__ import annotations

from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from tutorials.models import TutorialRegistration, TutorialSessions

HEADER = ['Title', 'First Name', 'Last Name', 'Student Ref', 'Email', 'Company', 'Attendance']
STATUS_LIST_FORMULA = '"ATTENDED,ABSENT,LATE,OTHER"'


def generate_roster_xlsx(session: TutorialSessions) -> bytes:
    """Return the xlsx as bytes (suitable for an email attachment)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Roster'

    # Header.
    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'

    # Body.
    registrations = (
        TutorialRegistration.objects
        .filter(tutorial_session=session)
        .select_related('student__user')
        .order_by('student__user__last_name', 'student__user__first_name')
    )
    for reg in registrations:
        user = reg.student.user
        ws.append([
            '',  # Title placeholder
            user.first_name or '',
            user.last_name or '',
            reg.student.student_ref,
            user.email or '',
            None,  # Company placeholder
            None,  # Attendance left blank for tutor input
        ])

    # Attendance dropdown — applied to all data rows.
    last_row = ws.max_row
    if last_row >= 2:
        dv = DataValidation(
            type='list', formula1=STATUS_LIST_FORMULA, allow_blank=True,
        )
        dv.error = 'Use one of: ATTENDED, ABSENT, LATE, OTHER.'
        dv.errorTitle = 'Invalid status'
        ws.add_data_validation(dv)
        attendance_col = get_column_letter(HEADER.index('Attendance') + 1)
        dv.add(f'{attendance_col}2:{attendance_col}{last_row}')

    # Sensible column widths.
    for idx, name in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, len(name) + 4)
    ws.column_dimensions['E'].width = 28  # Email
    ws.column_dimensions['G'].width = 14  # Attendance dropdown

    # Meta sheet (informational).
    meta = wb.create_sheet('Meta')
    event_code = session.tutorial_event.code if session.tutorial_event_id else ''
    meta.append(['session_id', session.id])
    meta.append(['event_code', event_code])
    meta.append(['session_date', str(session.start_date)])
    meta.append(['generated_at', timezone.now().isoformat()])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
