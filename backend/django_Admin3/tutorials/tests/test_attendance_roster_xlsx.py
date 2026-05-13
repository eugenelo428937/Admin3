"""Tests for the xlsx roster generator."""
from io import BytesIO

import openpyxl
from django.contrib.auth.models import User
from django.test import TestCase

from students.models import Student
from tutorials.models import TutorialRegistration
from tutorials.services.attendance_roster_xlsx import generate_roster_xlsx
from tutorials.tests.factories import make_event, make_session


class RosterXlsxTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.event = make_event(code='UT-XLS-1')
        cls.session = make_session(event=cls.event, title='Tutorial 1')
        u = User.objects.create_user(
            username='studentA', first_name='Alice', last_name='Aardvark',
            email='alice@example.com',
        )
        cls.student = Student.objects.create(user=u)
        TutorialRegistration.objects.create(
            student=cls.student, tutorial_session=cls.session,
        )

    def _load(self):
        return openpyxl.load_workbook(BytesIO(generate_roster_xlsx(self.session)))

    def test_returns_bytes(self):
        out = generate_roster_xlsx(self.session)
        self.assertIsInstance(out, bytes)
        self.assertGreater(len(out), 100)

    def test_header_row_columns(self):
        wb = self._load()
        ws = wb.active
        header = [c.value for c in ws[1]]
        self.assertEqual(
            header,
            ['Title', 'First Name', 'Last Name', 'Student Ref', 'Email', 'Company', 'Attendance'],
        )

    def test_first_data_row_populated(self):
        wb = self._load()
        ws = wb.active
        row2 = [c.value for c in ws[2]]
        self.assertEqual(row2[1], 'Alice')
        self.assertEqual(row2[2], 'Aardvark')
        self.assertEqual(row2[3], self.student.student_ref)
        self.assertEqual(row2[4], 'alice@example.com')
        self.assertEqual(row2[5], None)  # Company placeholder
        self.assertEqual(row2[6], None)  # Attendance left blank

    def test_attendance_dropdown_validation_present(self):
        wb = self._load()
        ws = wb.active
        dvs = list(ws.data_validations.dataValidation)
        self.assertTrue(dvs, 'expected at least one DataValidation on the sheet')
        statuses = '"ATTENDED,ABSENT,LATE,OTHER"'
        found = any(dv.formula1 == statuses and dv.type == 'list' for dv in dvs)
        self.assertTrue(found, f'attendance dropdown validation not found in {dvs!r}')

    def test_header_is_bold_and_freeze_panes_set(self):
        wb = self._load()
        ws = wb.active
        self.assertTrue(ws['A1'].font.bold)
        self.assertEqual(ws.freeze_panes, 'A2')

    def test_meta_sheet_exists_with_session_id(self):
        wb = self._load()
        self.assertIn('Meta', wb.sheetnames)
        meta = wb['Meta']
        values = {row[0].value: row[1].value for row in meta.iter_rows(min_row=1, max_row=4)}
        self.assertEqual(values.get('session_id'), self.session.id)
        self.assertEqual(values.get('event_code'), 'UT-XLS-1')
